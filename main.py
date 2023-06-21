# Import the required libraries
import cv2
from pyzbar.pyzbar import decode
from openpyxl import load_workbook

# Initialize the camera
cap = cv2.VideoCapture(0)

# Load the Excel database
workbook = load_workbook('database.xlsx')
sheet = workbook.active

# Get the column index for the code
code_column = 'B'

while True:
    # Capture a frame from the camera
    ret, frame = cap.read()

    # Decode the QR code from the captured frame
    qr_codes = decode(frame)

    # Check if any QR code is detected
    if qr_codes:
        # Extract the decoded data
        qr_data = qr_codes[0].data.decode('utf-8')
        print("QR Code Data:", qr_data)

        # Search for the QR code as a number in the Excel database
        code = None  # Initialize code as None
        for row in sheet.iter_rows(values_only=True):
            if str(qr_data) in map(str, row):
                row_index = row[0]
                code_cell = code_column + str(row_index)
                try:
                    code = sheet[code_cell].value
                except Exception as e:
                    print(f"Error accessing cell {code_cell}: {e}")
                else:
                    break
        
        if code is not None:
            print("hhehehehehe The member code is:", code)
        else:
            print("The member is not found.")

    # Display the frame
    cv2.imshow('QR Code Scanner', frame)

    # Check for the 'q' key to exit the loop
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the camera and close the window
cap.release()
cv2.destroyAllWindows()
